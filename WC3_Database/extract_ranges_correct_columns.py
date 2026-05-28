import openpyxl
import json

# Read Excel file
excel_path = r'H:\Pelit\PotS_JASS\WC3_ItemConcept\POTS_ItemConcept.xlsx'
wb = openpyxl.load_workbook(excel_path, data_only=True)
ws = wb['items_2']

print("=" * 100)
print("EXTRACTING ITEM LEVEL RANGES - CORRECT COLUMNS")
print("=" * 100)

# Correct column positions based on Row 10 (Chest) analysis:
# Col 4: Common range (e.g., "300-309")
# Col 11: Uncommon range (e.g., "310-319")
# Col 18: Rare range (e.g., "320-329")
# Col 25: Epic range (e.g., "330-339")
# Col 32: Legendary range (e.g., "340-349")

item_ranges = []

for row_idx in range(4, 26):  # Rows 4-25
    item_type = ws.cell(row_idx, 1).value
    base_level = ws.cell(row_idx, 2).value
    
    if not item_type or not base_level:
        continue
    
    # Extract ranges from correct columns
    common_range = ws.cell(row_idx, 4).value  # Column D (4)
    uncommon_range = ws.cell(row_idx, 11).value  # Column K (11)
    rare_range = ws.cell(row_idx, 18).value  # Column R (18)
    epic_range = ws.cell(row_idx, 25).value  # Column Y (25)
    legendary_range = ws.cell(row_idx, 32).value  # Column AF (32)
    
    print(f"\n{item_type} (Base: {base_level}):")
    
    ranges = {}
    
    # Parse Common range
    if common_range:
        try:
            if '-' in str(common_range):
                min_val, max_val = str(common_range).split('-')
                ranges['Common'] = {'min': int(min_val), 'max': int(max_val)}
                print(f"  Common    : {min_val}-{max_val}")
            else:
                # Single value - create 10-level range
                base = int(common_range)
                ranges['Common'] = {'min': base, 'max': base + 9}
                print(f"  Common    : {base}-{base+9}")
        except (ValueError, TypeError):
            print(f"  Common    : Skipping (non-numeric: {common_range})")
    
    # Parse Uncommon range
    if uncommon_range:
        try:
            if '-' in str(uncommon_range):
                min_val, max_val = str(uncommon_range).split('-')
                ranges['Uncommon'] = {'min': int(min_val), 'max': int(max_val)}
                print(f"  Uncommon  : {min_val}-{max_val}")
            else:
                base = int(uncommon_range)
                ranges['Uncommon'] = {'min': base, 'max': base + 9}
                print(f"  Uncommon  : {base}-{base+9}")
        except (ValueError, TypeError):
            print(f"  Uncommon  : Skipping (non-numeric: {uncommon_range})")
    
    # Parse Rare range
    if rare_range:
        try:
            if '-' in str(rare_range):
                min_val, max_val = str(rare_range).split('-')
                ranges['Rare'] = {'min': int(min_val), 'max': int(max_val)}
                print(f"  Rare      : {min_val}-{max_val}")
            else:
                base = int(rare_range)
                ranges['Rare'] = {'min': base, 'max': base + 9}
                print(f"  Rare      : {base}-{base+9}")
        except (ValueError, TypeError):
            print(f"  Rare      : Skipping (non-numeric: {rare_range})")
    
    # Parse Epic range
    if epic_range:
        try:
            if '-' in str(epic_range):
                min_val, max_val = str(epic_range).split('-')
                ranges['Epic'] = {'min': int(min_val), 'max': int(max_val)}
                print(f"  Epic      : {min_val}-{max_val}")
            else:
                base = int(epic_range)
                ranges['Epic'] = {'min': base, 'max': base + 9}
                print(f"  Epic      : {base}-{base+9}")
        except (ValueError, TypeError):
            print(f"  Epic      : Skipping (non-numeric: {epic_range})")
    
    # Parse Legendary range
    if legendary_range:
        try:
            if '-' in str(legendary_range):
                min_val, max_val = str(legendary_range).split('-')
                ranges['Legendary'] = {'min': int(min_val), 'max': int(max_val)}
                print(f"  Legendary : {min_val}-{max_val}")
            else:
                base = int(legendary_range)
                ranges['Legendary'] = {'min': base, 'max': base + 9}
                print(f"  Legendary : {base}-{base+9}")
        except (ValueError, TypeError):
            print(f"  Legendary : Skipping (non-numeric: {legendary_range})")
    
    item_ranges.append({
        'item_type': str(item_type).strip(),
        'base_level': str(base_level).strip() if isinstance(base_level, str) else base_level,
        'ranges': ranges
    })

wb.close()

# Save to JSON
output_path = r'H:\Pelit\PotS_JASS\WC3_Database\itemlevel_ranges_correct.json'
with open(output_path, 'w') as f:
    json.dump(item_ranges, f, indent=2)

print(f"\n\n{'='*100}")
print(f"Saved to: {output_path}")
print(f"Total item types with ranges: {len(item_ranges)}")
