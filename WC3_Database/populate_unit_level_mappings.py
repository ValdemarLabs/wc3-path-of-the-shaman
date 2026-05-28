import psycopg2
import openpyxl

try:
    conn = psycopg2.connect(
        host="localhost",
        database="wc3_pots",
        user="postgres",
        password="009900"
    )
    cursor = conn.cursor()
    
    print("=" * 100)
    print("POPULATING UNIT LEVEL MAPPINGS")
    print("=" * 100)
    
    # Create table
    with open(r'H:\Pelit\PotS_JASS\WC3_Database\create_unit_level_mappings.sql', 'r') as f:
        cursor.execute(f.read())
    conn.commit()
    
    # Read Excel
    excel_path = r'H:\Pelit\PotS_JASS\WC3_ItemConcept\POTS_ItemConcept.xlsx'
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb['items_2']
    
    # Map Excel names to database names
    excel_to_db = {
        'Helm': 'Head Armor',
        'Chest': 'Chest Armor',
        'Legpiece': 'Leg Armor',
        'Boots': 'Foot Armor',
        'Gloves': 'Hand Armor',
        'Rings': 'Ring',
        '2h': 'Two-Hand Weapon',
        'Back': 'Back',
        'Trinket': 'Trinket',
        'Shoulders': 'Shoulders',
        'Neck': 'Neck',
        'Bracers': 'Bracers',
        'Belt': 'Belt',
        '1h': '1h',
        'Stave': 'Stave',
        'Shield': 'Shield',
        'Miscellaneous': 'Miscellaneous',
    }
    
    # Rarity column mapping
    rarities = {
        'Common': (5, 10),      # columns 5-10
        'Uncommon': (12, 17),   # columns 12-17
        'Rare': (19, 24),       # columns 19-24
        'Epic': (26, 31),       # columns 26-31
        'Legendary': (33, 38),  # columns 33-38
    }
    
    print("\nInserting mappings:")
    insert_count = 0
    
    # Process each item type (rows 4-25)
    for row_idx in range(4, 26):
        item_type_excel = ws.cell(row_idx, 1).value
        if not item_type_excel:
            continue
        
        # Map to database name
        item_type_db = excel_to_db.get(str(item_type_excel).strip(), None)
        if not item_type_db:
            continue
        
        print(f"\n{item_type_excel} -> {item_type_db}:")
        
        # Process each rarity
        for rarity, (start_col, end_col) in rarities.items():
            # Get unit level ranges from row 3
            for col_idx in range(start_col, end_col + 1):
                unit_level_range = ws.cell(3, col_idx).value
                item_level = ws.cell(row_idx, col_idx).value
                
                if unit_level_range and item_level:
                    try:
                        item_level_int = int(item_level)
                        cursor.execute("""
                            INSERT INTO unit_level_mappings 
                            (item_class_name, rarity_name, unit_level_range, item_level)
                            VALUES (%s, %s, %s, %s)
                            ON CONFLICT (item_class_name, rarity_name, unit_level_range) 
                            DO UPDATE SET item_level = EXCLUDED.item_level
                        """, (item_type_db, rarity, unit_level_range, item_level_int))
                        insert_count += 1
                        print(f"  {rarity:10s} {unit_level_range:15s} -> iLvl {item_level_int}")
                    except (ValueError, TypeError):
                        pass
    
    conn.commit()
    wb.close()
    
    # Verify
    cursor.execute("SELECT COUNT(*) FROM unit_level_mappings")
    total = cursor.fetchone()[0]
    
    print(f"\n{'='*100}")
    print(f"✓ Inserted {insert_count} mappings")
    print(f"✓ Total mappings in database: {total}")
    print(f"{'='*100}")
    
    cursor.close()
    conn.close()
    
except Exception as e:
    print(f"ERROR: {e}")
    import traceback
    traceback.print_exc()
