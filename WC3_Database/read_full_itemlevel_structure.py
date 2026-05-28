import openpyxl

# Read Excel file
excel_path = r'H:\Pelit\PotS_JASS\WC3_ItemConcept\POTS_ItemConcept.xlsx'
wb = openpyxl.load_workbook(excel_path, data_only=True)
ws = wb['items_2']

print("=" * 100)
print("ITEMLEVEL RANGE STRUCTURE")
print("=" * 100)

# Read header rows (rows 1-3)
print("\nHeader structure:")
print("-" * 100)
for row_idx in range(1, 4):
    row_data = [str(ws.cell(row_idx, col).value or "") for col in range(1, 51)]
    print(f"Row {row_idx}: {row_data[:20]}")  # First 20 columns

print("\n" + "=" * 100)
print("ITEM TYPE -> RARITY RANGES")
print("=" * 100)

# Parse the structure - row 2 should have headers like "Common", "Uncommon", etc.
row2 = [ws.cell(2, col).value for col in range(1, 51)]
print(f"\nRow 2 headers: {[v for v in row2 if v]}")

# Identify column positions for each rarity
rarity_columns = {}
for col_idx, value in enumerate(row2, start=1):
    if value in ['Common', 'Uncommon', 'Rare', 'Epic', 'Legendary', 'Artifact']:
        rarity_columns[value] = col_idx

print(f"\nRarity column positions: {rarity_columns}")

print("\nParsing item type ranges:")
print("-" * 100)

# Parse data rows (starting from row 4, after headers)
item_ranges = []
for row_idx in range(4, 26):  # Rows 4-25 contain data
    item_type = ws.cell(row_idx, 1).value  # Column A
    base_level = ws.cell(row_idx, 2).value  # Column B
    
    if item_type and base_level:
        print(f"\n{item_type} (Base: {base_level}):")
        
        # For each rarity, find the range column
        ranges = {}
        for rarity, col_start in rarity_columns.items():
            # The range should be in the first column after rarity name (e.g., "300-309")
            range_value = ws.cell(row_idx, col_start + 1).value  # One column after rarity name
            if range_value:
                ranges[rarity] = range_value
                print(f"  {rarity}: {range_value}")
        
        item_ranges.append({
            'item_type': item_type,
            'base_level': base_level,
            'ranges': ranges
        })

print("\n" + "=" * 100)
print("SUMMARY FOR DATABASE")
print("=" * 100)

for item_data in item_ranges:
    print(f"\n{item_data['item_type']} (Base: {item_data['base_level']}):")
    for rarity, range_str in item_data['ranges'].items():
        if '-' in str(range_str):
            min_level, max_level = str(range_str).split('-')
            print(f"  {rarity:12s}: {min_level:4s} - {max_level:4s}")

wb.close()
