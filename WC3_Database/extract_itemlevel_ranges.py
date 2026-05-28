import openpyxl

# Read Excel file
excel_path = r'H:\Pelit\PotS_JASS\WC3_ItemConcept\POTS_ItemConcept.xlsx'
wb = openpyxl.load_workbook(excel_path, data_only=True)
ws = wb['items_2']

print("=" * 100)
print("PRECISE COLUMN MAPPING")
print("=" * 100)

# Map specific columns
print("\nAnalyzing Row 2 (headers) columns 1-50:")
for col_idx in range(1, 51):
    value = ws.cell(2, col_idx).value
    if value and isinstance(value, str):
        print(f"Col {col_idx:2d}: '{value}'")

print("\n" + "=" * 100)
print("ANALYZING CHEST ROW (Row 10) TO UNDERSTAND STRUCTURE")
print("=" * 100)

# Row 10 is Chest - use it as example
row_idx = 10
print(f"\nRow {row_idx} (Chest) - first 40 columns:")
for col_idx in range(1, 41):
    value = ws.cell(row_idx, col_idx).value
    header = ws.cell(2, col_idx).value
    print(f"Col {col_idx:2d} [{header}]: {value}")

print("\n" + "=" * 100)
print("EXTRACTING ITEM LEVEL RANGES")
print("=" * 100)

# Based on observation: 
# Column mapping appears to be:
# Col 1: Item type
# Col 2: Base level
# Col 3: Help text
# Col 4: "Common" (header)
# Col 5: Common range (e.g., "300-309")
# Col 11: "Uncommon" (header)  
# Col 12: Uncommon range (e.g., "310-319")
# Col 18: "Rare" (header)
# Col 19: Rare range
# Col 25: "Epic" (header)
# Col 26: Epic range
# Col 32: "Legendary" (header)
# Col 33: Legendary range

# Let's verify this with multiple rows
print("\nVerifying range columns for all item types:")
print("-" * 100)

item_ranges = []
for row_idx in range(4, 26):  # Rows 4-25
    item_type = ws.cell(row_idx, 1).value
    base_level = ws.cell(row_idx, 2).value
    
    if not item_type or not base_level:
        continue
    
    # Extract ranges from specific columns
    common_range = ws.cell(row_idx, 5).value  # Column E (5)
    uncommon_range = ws.cell(row_idx, 12).value  # Approximately col 12
    rare_range = ws.cell(row_idx, 19).value  # Approximately col 19
    epic_range = ws.cell(row_idx, 26).value  # Approximately col 26
    legendary_range = ws.cell(row_idx, 33).value  # Approximately col 33
    
    print(f"\n{item_type} (Base: {base_level}):")
    if common_range:
        print(f"  Common    : {common_range}")
    if uncommon_range:
        print(f"  Uncommon  : {uncommon_range}")
    if rare_range:
        print(f"  Rare      : {rare_range}")
    if epic_range:
        print(f"  Epic      : {epic_range}")
    if legendary_range:
        print(f"  Legendary : {legendary_range}")
    
    ranges = {}
    if common_range:
        ranges['Common'] = str(common_range)
    if uncommon_range:
        ranges['Uncommon'] = str(uncommon_range)
    if rare_range:
        ranges['Rare'] = str(rare_range)
    if epic_range:
        ranges['Epic'] = str(epic_range)
    if legendary_range:
        ranges['Legendary'] = str(legendary_range)
    
    item_ranges.append({
        'item_type': str(item_type).strip(),
        'base_level': base_level,
        'ranges': ranges
    })

wb.close()

# Save to JSON for database import
import json
output_path = r'H:\Pelit\PotS_JASS\WC3_Database\itemlevel_ranges.json'
with open(output_path, 'w') as f:
    json.dump(item_ranges, f, indent=2)

print(f"\n\nSaved to: {output_path}")
print(f"Total item types: {len(item_ranges)}")
