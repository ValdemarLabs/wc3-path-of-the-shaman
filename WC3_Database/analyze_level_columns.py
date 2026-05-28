import openpyxl

# Read Excel file
excel_path = r'H:\Pelit\PotS_JASS\WC3_ItemConcept\POTS_ItemConcept.xlsx'
wb = openpyxl.load_workbook(excel_path, data_only=True)
ws = wb['items_2']

print("=" * 100)
print("LEVEL RANGE COLUMNS ANALYSIS")
print("=" * 100)

# Header rows
print("\nRow 2 (headers):")
row2 = []
for col_idx in range(1, 50):
    value = ws.cell(2, col_idx).value
    if value:
        row2.append((col_idx, str(value)))

for col_idx, value in row2:
    print(f"  Col {col_idx:2d}: {value}")

print("\nRow 3 (sub-headers):")
for col_idx, _ in row2:
    value = ws.cell(3, col_idx).value
    if value:
        print(f"  Col {col_idx:2d}: {value}")

# Check Chest row (row 10) for level columns
print("\n" + "=" * 100)
print("CHEST (Row 10) - LEVEL COLUMNS")
print("=" * 100)

print("\nColumns 5-10 (Common rarity):")
for col_idx in range(5, 11):
    header = ws.cell(2, col_idx).value
    subheader = ws.cell(3, col_idx).value
    value = ws.cell(10, col_idx).value
    print(f"  Col {col_idx:2d} [{header}] [{subheader}]: {value}")

print("\nColumns 12-17 (Uncommon rarity):")
for col_idx in range(12, 18):
    header = ws.cell(2, col_idx).value
    subheader = ws.cell(3, col_idx).value
    value = ws.cell(10, col_idx).value
    print(f"  Col {col_idx:2d} [{header}] [{subheader}]: {value}")

# Extract structure for all rarities
print("\n" + "=" * 100)
print("UNIT LEVEL -> ITEM LEVEL MAPPING STRUCTURE")
print("=" * 100)

rarities = {
    'Common': (5, 10),      # columns 5-10
    'Uncommon': (12, 17),   # columns 12-17
    'Rare': (19, 24),       # columns 19-24
    'Epic': (26, 31),       # columns 26-31
    'Legendary': (33, 38),  # columns 33-38
}

print("\nMappings for Chest Armor:")
for rarity, (start_col, end_col) in rarities.items():
    print(f"\n{rarity}:")
    for col_idx in range(start_col, end_col + 1):
        subheader = ws.cell(3, col_idx).value
        item_level = ws.cell(10, col_idx).value
        if subheader and item_level:
            print(f"  {subheader:15s} -> iLvl {item_level}")

wb.close()
