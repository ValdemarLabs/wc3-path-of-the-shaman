#!/usr/bin/env python3
"""Analyze item level classification from POTS_ItemConcept.xlsx"""

import sys
try:
    import openpyxl
except ImportError:
    print("Installing openpyxl...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl

# Load the Excel file
xlsx_path = r"H:\Pelit\PotS_JASS\WC3_ItemConcept\POTS_ItemConcept.xlsx"
print(f"Loading {xlsx_path}...")
wb = openpyxl.load_workbook(xlsx_path, data_only=True)

# List all sheets
print(f"\nAvailable sheets: {wb.sheetnames}")

# Check if items_2 sheet exists
if "items_2" not in wb.sheetnames:
    print(f"\nERROR: 'items_2' sheet not found!")
    print(f"Available sheets: {wb.sheetnames}")
    sys.exit(1)

# Load items_2 sheet
ws = wb["items_2"]
print(f"\nAnalyzing 'items_2' sheet...")
print(f"Dimensions: {ws.max_row} rows x {ws.max_column} columns")

# Read headers
headers = []
for col in range(1, ws.max_column + 1):
    cell_value = ws.cell(1, col).value
    headers.append(cell_value)
    
print(f"\nHeaders ({len(headers)} columns):")
for i, header in enumerate(headers, 1):
    print(f"  {i:2d}. {header}")

# Find item level related columns
level_cols = []
for i, h in enumerate(headers, 1):
    if h and ('level' in str(h).lower() or 'tier' in str(h).lower() or 'class' in str(h).lower() or 'rarity' in str(h).lower()):
        level_cols.append((i, h))

print(f"\nItem level/classification related columns:")
for col_idx, col_name in level_cols:
    print(f"  Column {col_idx}: {col_name}")

# Analyze item level values
print(f"\n{'='*80}")
print("ITEM LEVEL ANALYSIS")
print(f"{'='*80}")

# Find the item level column
itemlevel_col_idx = None
itemcode_col_idx = None
itemname_col_idx = None
rarity_col_idx = None

for i, h in enumerate(headers, 1):
    h_lower = str(h).lower() if h else ""
    if 'item level' in h_lower or 'itemlevel' in h_lower or h_lower == 'level':
        itemlevel_col_idx = i
    if 'item code' in h_lower or 'itemcode' in h_lower or h_lower == 'code':
        itemcode_col_idx = i
    if 'item name' in h_lower or 'itemname' in h_lower or h_lower == 'name':
        itemname_col_idx = i
    if 'rarity' in h_lower or 'tier' in h_lower:
        rarity_col_idx = i

print(f"\nKey column indices:")
print(f"  Item Code: Column {itemcode_col_idx} ({headers[itemcode_col_idx-1] if itemcode_col_idx else 'NOT FOUND'})")
print(f"  Item Name: Column {itemname_col_idx} ({headers[itemname_col_idx-1] if itemname_col_idx else 'NOT FOUND'})")
print(f"  Item Level: Column {itemlevel_col_idx} ({headers[itemlevel_col_idx-1] if itemlevel_col_idx else 'NOT FOUND'})")
print(f"  Rarity/Tier: Column {rarity_col_idx} ({headers[rarity_col_idx-1] if rarity_col_idx else 'NOT FOUND'})")

if itemlevel_col_idx:
    # Collect item level data
    itemlevel_data = {}
    for row in range(2, min(ws.max_row + 1, 1000)):  # Limit to first 1000 rows
        itemcode = ws.cell(row, itemcode_col_idx).value if itemcode_col_idx else None
        itemname = ws.cell(row, itemname_col_idx).value if itemname_col_idx else None
        itemlevel = ws.cell(row, itemlevel_col_idx).value
        rarity = ws.cell(row, rarity_col_idx).value if rarity_col_idx else None
        
        if itemlevel is not None:
            try:
                level_int = int(itemlevel) if itemlevel else 0
                if level_int not in itemlevel_data:
                    itemlevel_data[level_int] = []
                itemlevel_data[level_int].append({
                    'code': itemcode,
                    'name': itemname,
                    'level': level_int,
                    'rarity': rarity
                })
            except (ValueError, TypeError):
                pass
    
    # Sort levels
    sorted_levels = sorted(itemlevel_data.keys())
    
    print(f"\n{'='*80}")
    print("ITEM LEVEL DISTRIBUTION")
    print(f"{'='*80}")
    print(f"Total unique item levels: {len(sorted_levels)}")
    print(f"Level range: {min(sorted_levels)} to {max(sorted_levels)}")
    print(f"\nDistribution:")
    for level in sorted_levels[:50]:  # Show first 50 levels
        count = len(itemlevel_data[level])
        print(f"  Level {level:3d}: {count:3d} items")
    
    if len(sorted_levels) > 50:
        print(f"  ... and {len(sorted_levels) - 50} more levels")
    
    # Show examples for key level ranges
    print(f"\n{'='*80}")
    print("ITEM LEVEL EXAMPLES")
    print(f"{'='*80}")
    
    # Show examples from different level ranges
    ranges = [
        (1, 10, "Very Low Level (1-10)"),
        (11, 50, "Low Level (11-50)"),
        (51, 100, "Medium Level (51-100)"),
        (101, 500, "High Level (101-500)"),
        (501, 1000, "Very High Level (501-1000)"),
    ]
    
    for min_lvl, max_lvl, desc in ranges:
        matching_levels = [l for l in sorted_levels if min_lvl <= l <= max_lvl]
        if matching_levels:
            print(f"\n{desc}:")
            # Show a few examples
            sample_level = matching_levels[len(matching_levels)//2]  # Pick middle level
            items = itemlevel_data[sample_level][:3]  # Show 3 items
            for item in items:
                rarity_str = f" [{item['rarity']}]" if item['rarity'] else ""
                print(f"  Level {item['level']}: {item['code']} - {item['name']}{rarity_str}")
    
    # Analyze level clustering/grouping patterns
    print(f"\n{'='*80}")
    print("LEVEL CLUSTERING ANALYSIS")
    print(f"{'='*80}")
    
    # Try to identify level brackets/tiers
    level_counts = [(level, len(itemlevel_data[level])) for level in sorted_levels]
    
    # Find levels with multiple items (likely tier boundaries)
    popular_levels = [l for l, count in level_counts if count >= 5]
    print(f"\nPopular item levels (5+ items): {popular_levels[:20]}")
    
    # Check for patterns in level numbering
    print(f"\nChecking for tier patterns...")
    
    # Common tier patterns in WC3
    tier_patterns = {
        "Even numbers (2, 4, 6...)": [l for l in sorted_levels if l % 2 == 0 and l <= 50],
        "Multiples of 5 (5, 10, 15...)": [l for l in sorted_levels if l % 5 == 0 and l <= 100],
        "Multiples of 10 (10, 20, 30...)": [l for l in sorted_levels if l % 10 == 0 and l <= 200],
        "Multiples of 100 (100, 200, 300...)": [l for l in sorted_levels if l % 100 == 0 and l <= 1000],
    }
    
    for pattern_name, pattern_levels in tier_patterns.items():
        if pattern_levels:
            item_count = sum(len(itemlevel_data[l]) for l in pattern_levels)
            print(f"  {pattern_name}: {len(pattern_levels)} levels, {item_count} items")

else:
    print("\nERROR: Could not find item level column!")

print(f"\n{'='*80}")
print("RECOMMENDATIONS FOR DATABASE")
print(f"{'='*80}")
print("""
Based on the analysis above, the item level system should:

1. Use the exact item level values from the spreadsheet
2. Store item levels in the 'item_level' column (already exists in database)
3. Create level brackets/tiers if the drop system uses ranges
4. Consider adding an 'item_tier' or 'level_bracket' column if needed

To update the database with correct item levels:
1. Match items by item_code
2. Update the item_level column with values from the spreadsheet
3. Ensure drop system queries use: WHERE item_level BETWEEN X AND Y
""")
