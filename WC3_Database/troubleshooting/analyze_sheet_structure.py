#!/usr/bin/env python3
"""Deep dive into items_2 sheet structure"""

import openpyxl

xlsx_path = r"H:\Pelit\PotS_JASS\WC3_ItemConcept\POTS_ItemConcept.xlsx"
wb = openpyxl.load_workbook(xlsx_path, data_only=True)
ws = wb["items_2"]

print("="*100)
print("FULL SHEET CONTENT (First 30 rows)")
print("="*100)

# Look at all content
for row_idx in range(1, min(31, ws.max_row + 1)):
    print(f"\nRow {row_idx}:")
    row_data = []
    for col_idx in range(1, min(51, ws.max_column + 1)):
        cell_value = ws.cell(row_idx, col_idx).value
        if cell_value is not None:
            row_data.append(f"  Col {col_idx}: {cell_value}")
    
    if row_data:
        print("\n".join(row_data))
    else:
        print("  (empty row)")

print("\n" + "="*100)
print("LOOKING FOR ITEM LEVEL PATTERNS")
print("="*100)

# Search for cells that look like item levels
print("\nSearching for numeric values that could be item levels...")
item_level_candidates = []

for row_idx in range(1, ws.max_row + 1):
    for col_idx in range(1, ws.max_column + 1):
        cell_value = ws.cell(row_idx, col_idx).value
        
        # Check if this looks like an item level (numeric value)
        if isinstance(cell_value, (int, float)) and cell_value > 0 and cell_value < 10000:
            # Check nearby cells for context
            above = ws.cell(row_idx - 1, col_idx).value if row_idx > 1 else None
            left = ws.cell(row_idx, col_idx - 1).value if col_idx > 1 else None
            below = ws.cell(row_idx + 1, col_idx).value if row_idx < ws.max_row else None
            
            item_level_candidates.append({
                'row': row_idx,
                'col': col_idx,
                'value': cell_value,
                'above': above,
                'left': left,
                'below': below
            })

print(f"\nFound {len(item_level_candidates)} numeric values")
print("\nFirst 50 candidates:")
for candidate in item_level_candidates[:50]:
    print(f"  [{candidate['row']}, {candidate['col']}] = {candidate['value']}")
    if candidate['above']:
        print(f"    ↑ Above: {candidate['above']}")
    if candidate['left']:
        print(f"    ← Left: {candidate['left']}")
    if candidate['below']:
        print(f"    ↓ Below: {candidate['below']}")

# Look for patterns
print("\n" + "="*100)
print("LEVEL PATTERN ANALYSIS")
print("="*100)

# Group by column
from collections import defaultdict
by_column = defaultdict(list)
for c in item_level_candidates:
    by_column[c['col']].append(c['value'])

print("\nValues by column:")
for col in sorted(by_column.keys())[:20]:
    values = sorted(set(by_column[col]))
    if len(values) > 2:  # Only show columns with multiple different values
        print(f"  Column {col}: {values}")

# Check if these are level brackets
print("\nAnalyzing level progression patterns...")
for col in sorted(by_column.keys())[:10]:
    values = sorted(by_column[col])
    if len(values) >= 3:
        diffs = [values[i+1] - values[i] for i in range(len(values)-1)]
        if diffs:
            avg_diff = sum(diffs) / len(diffs)
            print(f"\nColumn {col}:")
            print(f"  Values: {values[:10]}")
            print(f"  Differences: {diffs[:10]}")
            print(f"  Average step: {avg_diff:.1f}")
            
            # Check if constant progression
            if all(abs(d - avg_diff) < 10 for d in diffs):
                print(f"  → Regular progression! Steps of ~{avg_diff:.0f}")
