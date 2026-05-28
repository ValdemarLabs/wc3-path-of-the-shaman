#!/usr/bin/env python3
"""Extract complete item level classification system from POTS_ItemConcept.xlsx"""

import openpyxl

xlsx_path = r"H:\Pelit\PotS_JASS\WC3_ItemConcept\POTS_ItemConcept.xlsx"
wb = openpyxl.load_workbook(xlsx_path, data_only=True)
ws = wb["items_2"]

print("="*100)
print("ITEM LEVEL CLASSIFICATION SYSTEM")
print("="*100)

# Parse the header structure from rows 1-3
print("\nRARITY TIERS:")
print("-" * 100)

# Rarity columns are in row 2
rarities = {}
rarity_mappings = {
    4: "Common",
    11: "Uncommon", 
    18: "Rare",
    25: "Epic",
    32: "Legendary"
}

for col_idx, rarity_name in rarity_mappings.items():
    # Get the base item level for this rarity
    base_level = ws.cell(2, col_idx + 1).value
    
    # Get the level ranges from row 3
    level_ranges = []
    for i in range(6):  # 6 sub-levels per rarity
        col = col_idx + 1 + i
        level_val = ws.cell(2, col).value
        level_desc = ws.cell(3, col).value
        if level_val is not None:
            level_ranges.append((level_val, level_desc))
    
    rarities[rarity_name] = {
        'base_level': base_level,
        'ranges': level_ranges
    }
    
    print(f"\n{rarity_name}:")
    print(f"  Base Level: {base_level}")
    print(f"  Progression: {[r[0] for r in level_ranges]}")

# Parse item type classifications from rows 4 onwards
print("\n" + "="*100)
print("ITEM TYPE LEVEL RANGES")
print("="*100)

item_types = []

for row_idx in range(4, ws.max_row + 1):
    item_type = ws.cell(row_idx, 1).value
    base_level = ws.cell(row_idx, 2).value
    help_text = ws.cell(row_idx, 3).value
    
    if item_type and base_level is not None:
        # Get item classification flags
        is_permanent = ws.cell(row_idx, 41).value
        is_charged = ws.cell(row_idx, 42).value
        is_powerup = ws.cell(row_idx, 43).value
        is_artifact = ws.cell(row_idx, 44).value
        is_purchasable = ws.cell(row_idx, 45).value
        is_campaign = ws.cell(row_idx, 46).value
        is_misc = ws.cell(row_idx, 47).value
        
        # Get level ranges for each rarity
        level_data = {}
        for rarity_name, rarity_col in [("Common", 4), ("Uncommon", 11), ("Rare", 18), ("Epic", 25), ("Legendary", 32)]:
            range_desc = ws.cell(row_idx, rarity_col).value
            levels = []
            for i in range(6):
                level_val = ws.cell(row_idx, rarity_col + 1 + i).value
                if level_val:
                    levels.append(level_val)
            if levels:
                level_data[rarity_name] = {
                    'range': range_desc,
                    'levels': levels,
                    'min': min(levels),
                    'max': max(levels)
                }
        
        item_types.append({
            'type': item_type,
            'base_level': base_level,
            'help': help_text,
            'level_data': level_data,
            'is_permanent': is_permanent,
            'is_charged': is_charged,
            'is_powerup': is_powerup,
            'is_artifact': is_artifact,
            'is_purchasable': is_purchasable,
            'is_campaign': is_campaign,
            'is_misc': is_misc
        })

# Display item types with their level ranges
for item_type_data in item_types:
    print(f"\n{item_type_data['type']} (Base Level: {item_type_data['base_level']})")
    print(f"  Description: {item_type_data['help']}")
    
    # Classification
    classifications = []
    if item_type_data['is_permanent']: classifications.append("Permanent")
    if item_type_data['is_charged']: classifications.append("Charged")
    if item_type_data['is_powerup']: classifications.append("Power Up")
    if item_type_data['is_artifact']: classifications.append("Artifact")
    if item_type_data['is_purchasable']: classifications.append("Purchasable")
    if item_type_data['is_campaign']: classifications.append("Campaign")
    if item_type_data['is_misc']: classifications.append("Miscellaneous")
    
    if classifications:
        print(f"  Classification: {', '.join(classifications)}")
    
    # Level ranges by rarity
    if item_type_data['level_data']:
        print(f"  Level Ranges by Rarity:")
        for rarity in ["Common", "Uncommon", "Rare", "Epic", "Legendary"]:
            if rarity in item_type_data['level_data']:
                rd = item_type_data['level_data'][rarity]
                print(f"    {rarity:12s}: {rd['range']:10s} → levels {rd['min']}-{rd['max']} ({len(rd['levels'])} values)")

# Create SQL-friendly level bracket summary
print("\n" + "="*100)
print("LEVEL BRACKET SUMMARY FOR DROP SYSTEM")
print("="*100)

print("""
This system works as follows:

1. Item Slots are assigned base level ranges (100-149, 150-199, etc.)
2. Within each slot, there are 5 rarity tiers
3. Each rarity has 6 progression levels

For the drop system, you should:
""")

for item_type_data in item_types:
    if item_type_data['level_data']:
        min_level = min(rd['min'] for rd in item_type_data['level_data'].values())
        max_level = max(rd['max'] for rd in item_type_data['level_data'].values())
        print(f"\n{item_type_data['type']:15s}: Levels {min_level:3d}-{max_level:3d}")
        
        for rarity in ["Common", "Uncommon", "Rare", "Epic", "Legendary"]:
            if rarity in item_type_data['level_data']:
                rd = item_type_data['level_data'][rarity]
                print(f"  {rarity:12s}: {rd['min']:3d}-{rd['max']:3d}")

# Generate SQL for database update
print("\n" + "="*100)
print("DATABASE SCHEMA RECOMMENDATIONS")
print("="*100)

print("""
To support this item level system in the database:

1. Keep the existing 'item_level' column (stores exact level like 105, 132, etc.)

2. Add helper columns for easier queries:
   ALTER TABLE items ADD COLUMN item_slot VARCHAR(50);
   ALTER TABLE items ADD COLUMN level_bracket VARCHAR(20);
   
3. Create a lookup table for level brackets:
   CREATE TABLE item_level_brackets (
       id SERIAL PRIMARY KEY,
       slot_name VARCHAR(50),
       rarity_name VARCHAR(20),
       min_level INTEGER,
       max_level INTEGER,
       base_level INTEGER
   );

4. Item level brackets should match your drop system:
""")

print("\nINSERT statements for item_level_brackets table:")
print("INSERT INTO item_level_brackets (slot_name, rarity_name, min_level, max_level, base_level) VALUES")

insert_values = []
for item_type_data in item_types:
    if item_type_data['level_data']:
        for rarity in ["Common", "Uncommon", "Rare", "Epic", "Legendary"]:
            if rarity in item_type_data['level_data']:
                rd = item_type_data['level_data'][rarity]
                insert_values.append(
                    f"  ('{item_type_data['type']}', '{rarity}', {rd['min']}, {rd['max']}, {item_type_data['base_level']})"
                )

print(",\n".join(insert_values) + ";")

print("\n" + "="*100)
print("CURRENT DATABASE STATUS CHECK")
print("="*100)
print("""
Run this SQL to see how imported item levels match the classification system:

SELECT 
    CASE 
        WHEN item_level BETWEEN 0 AND 49 THEN 'Other (0-49)'
        WHEN item_level BETWEEN 50 AND 99 THEN 'Miscellaneous (50-99)'
        WHEN item_level BETWEEN 100 AND 149 THEN 'Helm (100-149)'
        WHEN item_level BETWEEN 150 AND 199 THEN 'Neck (150-199)'
        WHEN item_level BETWEEN 200 AND 249 THEN 'Shoulder (200-249)'
        WHEN item_level BETWEEN 250 AND 299 THEN 'Back (250-299)'
        -- Add more ranges based on output above
        ELSE 'Unknown'
    END as level_bracket,
    COUNT(*) as item_count,
    MIN(item_level) as min_level,
    MAX(item_level) as max_level
FROM items
WHERE item_level > 0
GROUP BY level_bracket
ORDER BY MIN(item_level);
""")
