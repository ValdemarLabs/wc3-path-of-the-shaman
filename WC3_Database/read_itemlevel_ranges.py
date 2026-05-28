import openpyxl
import psycopg2

# Read Excel file
excel_path = r'H:\Pelit\PotS_JASS\WC3_ItemConcept\POTS_ItemConcept.xlsx'
wb = openpyxl.load_workbook(excel_path, data_only=True)

print("=" * 80)
print("ANALYZING ITEMLEVEL RANGES FROM EXCEL")
print("=" * 80)

# Check sheet "items_2"
if 'items_2' in wb.sheetnames:
    ws = wb['items_2']
    print(f"\nSheet 'items_2' found. Dimensions: {ws.dimensions}")
    print("\nFirst 30 rows:")
    print("-" * 80)
    
    for row_idx in range(1, min(31, ws.max_row + 1)):
        row_data = []
        for col_idx in range(1, min(15, ws.max_column + 1)):  # First 15 columns
            cell = ws.cell(row_idx, col_idx)
            value = cell.value
            if value is not None:
                row_data.append(str(value)[:30])  # Truncate long values
            else:
                row_data.append("")
        
        if any(row_data):  # Only print non-empty rows
            print(f"Row {row_idx:2d}: {' | '.join(row_data)}")
else:
    print("\nSheet 'items_2' not found. Available sheets:")
    for sheet_name in wb.sheetnames:
        print(f"  - {sheet_name}")

print("\n" + "=" * 80)
print("CURRENT DATABASE CONSTRAINTS")
print("=" * 80)

try:
    conn = psycopg2.connect(
        host="localhost",
        database="wc3_pots",
        user="postgres",
        password="009900"
    )
    cursor = conn.cursor()
    
    # Check current check constraints on items table
    cursor.execute("""
        SELECT conname, pg_get_constraintdef(oid) 
        FROM pg_constraint 
        WHERE conrelid = 'items'::regclass 
        AND contype = 'c'
        ORDER BY conname
    """)
    
    constraints = cursor.fetchall()
    print("\nCheck constraints on 'items' table:")
    for name, definition in constraints:
        print(f"  {name}: {definition}")
    
    cursor.close()
    conn.close()
    
except Exception as e:
    print(f"Database error: {e}")

wb.close()
